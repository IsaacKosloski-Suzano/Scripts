import pandas as pd
from openpyxl import load_workbook

def fill_empty_columns(input_file, output_file):
    # Load the workbook
    wb = load_workbook(input_file)
    if len(wb.sheetnames) < 2:
        raise ValueError("The Excel file must contain at least two sheets.")
    
    # Load both sheets into DataFrames
    df1 = pd.read_excel(input_file, sheet_name=wb.sheetnames[0])
    df2 = pd.read_excel(input_file, sheet_name=wb.sheetnames[1])
    
    # Identify columns in the first sheet that are empty (only have a title)
    empty_columns = [col for col in df1.columns if df1[col].dropna().empty]
    
    # Fill empty columns with data from the second sheet if the column exists
    for col in empty_columns:
        if col in df2.columns:
            df1[col] = df2[col]
    
    # Save the updated DataFrame back to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name=wb.sheetnames[0], index=False)
        df2.to_excel(writer, sheet_name=wb.sheetnames[1], index=False)
    
    print("Empty columns in the first sheet have been filled with data from the second sheet where available.")

# Example usage:
fill_empty_columns('data02.xlsx', 'output03.xlsx')
