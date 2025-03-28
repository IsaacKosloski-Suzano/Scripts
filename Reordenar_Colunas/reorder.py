import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def reorder_columns(input_file, output_file):
    # Load the workbook and sheets
    wb = load_workbook(input_file)
    sheet1 = wb[wb.sheetnames[0]]  # First sheet
    sheet2 = wb[wb.sheetnames[1]]  # Second sheet
    
    # Get column order from the first sheet
    col_order = [cell.value for cell in sheet1[1] if cell.value]
    
    # Load second sheet into DataFrame
    df2 = pd.read_excel(input_file, sheet_name=wb.sheetnames[1])
    
    # Identify existing columns in second sheet
    existing_cols = list(df2.columns)
    
    # Columns to keep in order
    ordered_cols = [col for col in col_order if col in existing_cols]
    
    # Columns that exist in second sheet but not in first
    extra_cols = [col for col in existing_cols if col not in col_order]
    
    # Columns that are in first sheet but missing in second
    missing_cols = [col for col in col_order if col not in existing_cols]
    
    # Create missing columns with NaN values
    for col in missing_cols:
        df2[col] = None
    
    # Reorder columns
    final_columns = ordered_cols + missing_cols + extra_cols
    df2 = df2[final_columns]
    
    # Save to new workbook
    df2.to_excel(output_file, sheet_name=wb.sheetnames[1], index=False)
    
    # Reload workbook to style missing column headers
    wb_output = load_workbook(output_file)
    ws_output = wb_output[wb.sheetnames[1]]
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    bold_font = Font(bold=True, color="FFFFFF")
    
    # Apply styles to missing column headers
    for col_idx, col_name in enumerate(final_columns, start=1):
        if col_name in missing_cols:
            cell = ws_output.cell(row=1, column=col_idx)
            cell.fill = red_fill
            cell.font = bold_font
    
    # Save final workbook
    wb_output.save(output_file)
    print("Columns reordered and missing columns highlighted in red.")

# Example usage:
reorder_columns('data.xlsx', 'output.xlsx')
